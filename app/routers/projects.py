import io
import json
import logging
from threading import Thread

from fastapi import APIRouter, Depends, Request, Form, HTTPException, Response, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse

logger = logging.getLogger("autoblogspot.billing")
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..database import get_db
from ..dependencies import get_current_user, can_create_project
from ..models import Project, ProjectSite, BlogspotSite, Keyword, KeywordCluster, Article
from ..services.openrouter import FREE_MODELS, LANGUAGE_NAMES, get_setting
from ..services.ai_providers import CLAUDE_MODELS, OPENAI_MODELS, GROQ_MODELS, GEMINI_MODELS
from ..services.scheduler import process_keyword_clustering
from ..templates import templates


def _get_available_models(db, user_id: int) -> dict:
    """Return model groups based on which API keys the user has configured."""
    groups = {"openrouter": FREE_MODELS, "gemini": [], "claude": [], "openai": [], "groq": []}
    gemini_keys = get_setting(db, "gemini_api_keys", user_id=user_id)
    if gemini_keys and gemini_keys.strip():
        groups["gemini"] = GEMINI_MODELS
    if get_setting(db, "claude_api_key", user_id=user_id):
        groups["claude"] = CLAUDE_MODELS
    if get_setting(db, "openai_api_key", user_id=user_id):
        groups["openai"] = OPENAI_MODELS
    if get_setting(db, "groq_api_key", user_id=user_id):
        groups["groq"] = GROQ_MODELS
    return groups


def _get_default_model(groups: dict, current_model: str = None) -> str:
    """Trả về model mặc định theo ưu tiên: Gemini → Claude → OpenAI → Groq → OpenRouter."""
    if current_model:
        return current_model  # giữ nguyên nếu đã chọn
    for provider in ("gemini", "claude", "openai", "groq"):
        if groups.get(provider):
            return groups[provider][0]["id"]
    or_models = groups.get("openrouter", FREE_MODELS)
    return or_models[1]["id"] if len(or_models) > 1 else "meta-llama/llama-3.3-70b-instruct:free"


router = APIRouter()
LANGUAGES = list(LANGUAGE_NAMES.items())


def _parse_project_form(backlinks_json: str, custom_labels_text: str) -> tuple[list, list]:
    try:
        backlinks = json.loads(backlinks_json) if backlinks_json else []
    except json.JSONDecodeError:
        backlinks = []
    custom_labels = [l.strip() for l in custom_labels_text.split(",") if l.strip()] if custom_labels_text.strip() else []
    return backlinks, custom_labels


def _cap_articles_per_day(value: int, user) -> int:
    sub = user.subscription
    if not sub or sub.articles_per_day_limit is None:
        return value
    return min(value, sub.articles_per_day_limit)


def _build_site_groups(sites: list) -> list[dict]:
    """Group sites by Google account or platform account for display."""
    seen: dict[str, dict] = {}
    result: list[dict] = []
    for site in sites:
        platform = site.platform or "blogspot"
        if platform == "blogspot":
            acct = site.account
            key = f"google_{acct.id}" if acct else "blogspot_other"
            if key not in seen:
                label = (acct.name or acct.email) if acct else "Blogspot"
                sub_label = acct.email if acct and acct.name else ""
                entry = {"label": label, "sub_label": sub_label, "icon": "blogspot", "sites": []}
                seen[key] = entry
                result.append(entry)
            seen[key]["sites"].append(site)
        else:
            acct = site.platform_account
            key = f"{platform}_{acct.id}" if acct else f"{platform}_other"
            if key not in seen:
                label = (acct.name or platform.title()) if acct else platform.title()
                entry = {"label": label, "sub_label": "", "icon": platform, "sites": []}
                seen[key] = entry
                result.append(entry)
            seen[key]["sites"].append(site)
    return result


@router.get("/projects", response_class=HTMLResponse)
def projects_page(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    projects = db.query(Project).filter(Project.user_id == current_user.id).order_by(Project.created_at.desc()).all()
    return templates.TemplateResponse(request, "projects.html", {
        "projects": projects, "current_user": current_user, "active_page": "projects",
    })


@router.get("/projects/new", response_class=HTMLResponse)
def new_project_page(request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    sites = db.query(BlogspotSite).filter(
        BlogspotSite.user_id == current_user.id, BlogspotSite.is_active == True
    ).all()
    model_groups = _get_available_models(db, current_user.id)
    return templates.TemplateResponse(request, "project_form.html", {
        "sites": sites, "site_groups": _build_site_groups(sites), "languages": LANGUAGES,
        "models": FREE_MODELS,
        "model_groups": model_groups,
        "default_model": _get_default_model(model_groups),
        "project": None, "current_user": current_user, "active_page": "projects",
    })


@router.post("/projects/new")
def create_project(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    articles_per_day: int = Form(3),
    min_interval: int = Form(60),
    max_interval: int = Form(240),
    ai_model: str = Form("meta-llama/llama-3.1-8b-instruct:free"),
    site_ids: list[int] = Form(default=[]),
    site_languages: list[str] = Form(default=[]),
    backlinks_json: str = Form("[]"),
    custom_labels_text: str = Form(""),
    keywords_text: str = Form(""),
    db: Session = Depends(get_db),
):
    current_user = get_current_user(request, db)
    ok, msg = can_create_project(current_user, db)
    if not ok:
        return RedirectResponse(f"/projects?error={msg}", status_code=303)

    backlinks, custom_labels = _parse_project_form(backlinks_json, custom_labels_text)
    project = Project(
        user_id=current_user.id,
        name=name, description=description,
        articles_per_day=_cap_articles_per_day(articles_per_day, current_user),
        min_interval_minutes=min_interval, max_interval_minutes=max_interval,
        ai_model=ai_model,
        backlinks=json.dumps(backlinks),
        custom_labels=json.dumps(custom_labels, ensure_ascii=False),
        status="pending",
    )
    db.add(project)
    db.flush()

    for i, site_id in enumerate(site_ids):
        lang = site_languages[i] if i < len(site_languages) else "vi"
        db.add(ProjectSite(project_id=project.id, site_id=site_id, language=lang))

    if keywords_text.strip():
        for kw in [k.strip() for k in keywords_text.splitlines() if k.strip()]:
            db.add(Keyword(project_id=project.id, keyword=kw))

    db.commit()
    return RedirectResponse(f"/projects/{project.id}", status_code=303)


@router.get("/projects/clusters-template")
def download_clusters_template():
    """Tải file Excel mẫu để nhập nhóm từ khóa đã chia sẵn."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhóm từ khóa"

    header_fill  = PatternFill("solid", fgColor="1D4ED8")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    example_fill = PatternFill("solid", fgColor="EFF6FF")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Tên nhóm bài viết (Cluster Name)", "Từ khóa 1", "Từ khóa 2", "Từ khóa 3", "Từ khóa 4", "Từ khóa 5", "Từ khóa 6"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    examples = [
        ["Ghế công viên giá rẻ tại HCM", "ghế công viên giá rẻ", "ghế băng công viên rẻ", "mua ghế công viên hcm", "ghế công viên gỗ giá rẻ", "", ""],
        ["Ghế inox ngoài trời bền đẹp",  "ghế inox ngoài trời", "ghế inox chịu mưa nắng", "ghế inox sân vườn", "bàn ghế inox ngoài trời", "ghế inox 304", ""],
        ["Cách chọn ghế công viên phù hợp", "cách chọn ghế công viên", "tiêu chí chọn ghế ngoài trời", "ghế công viên loại nào tốt", "", "", ""],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = example_fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 45
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28
    ws.row_dimensions[1].height = 35

    ws2 = wb.create_sheet("Hướng dẫn")
    guide = [
        ("AutoBlogspot — Hướng dẫn nhập nhóm từ khóa", True),
        ("", False),
        ("1. Mỗi hàng = 1 nhóm bài viết (cluster)", False),
        ("2. Cột A: Tên nhóm — đây sẽ là định hướng tiêu đề bài viết (bắt buộc)", False),
        ("3. Cột B trở đi: Các từ khóa thuộc nhóm đó (tối thiểu 1 từ khóa, tối đa không giới hạn)", False),
        ("4. Mỗi hàng có thể có số lượng từ khóa khác nhau — bỏ trống ô nếu không dùng", False),
        ("5. Không thay đổi hàng tiêu đề (hàng 1)", False),
        ("6. Có thể thêm cột từ khóa tùy ý (cột H, I, J... — không giới hạn)", False),
        ("", False),
        ("LỢI ÍCH: Upload file này thay vì để AI tự chia → chiến dịch bắt đầu ngay lập tức!", True),
    ]
    ws2.column_dimensions["A"].width = 80
    for i, (text, bold) in enumerate(guide, 1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font      = Font(bold=bold, size=11 if bold else 10)
        cell.alignment = Alignment(wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau-nhom-tu-khoa.xlsx"},
    )


@router.get("/projects/{project_id}", response_class=HTMLResponse)
def project_detail(project_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    clusters = db.query(KeywordCluster).filter(KeywordCluster.project_id == project_id).all()
    keywords = db.query(Keyword).filter(Keyword.project_id == project_id).all()

    rows = db.query(Article.status, func.count().label("n")).filter(
        Article.project_id == project_id
    ).group_by(Article.status).all()
    counts = {status: n for status, n in rows}

    recent_articles = (
        db.query(Article)
        .filter(Article.project_id == project_id, Article.status == "published")
        .order_by(Article.published_at.desc()).limit(10).all()
    )

    return templates.TemplateResponse(request, "project_detail.html", {
        "project": project, "clusters": clusters, "keywords": keywords,
        "total_articles":  sum(counts.values()),
        "published":       counts.get("published", 0),
        "pending_count":   counts.get("pending", 0),
        "writing_count":   counts.get("writing", 0),
        "failed_count":    counts.get("failed", 0),
        "recent_articles": recent_articles,
        "backlinks":       json.loads(project.backlinks or "[]"),
        "current_user":    current_user,
        "active_page":     "projects",
    })


def _trigger_pipeline():
    process_keyword_clustering()


@router.post("/projects/{project_id}/start")
def start_project(project_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404)
    # Check subscription still active
    sub = current_user.subscription
    if not sub or not sub.is_active_plan:
        return RedirectResponse(f"/projects/{project_id}?error=Goi+dich+vu+het+han.+Vui+long+gia+han.", status_code=303)
    project.status = "running"
    db.commit()
    Thread(target=_trigger_pipeline, daemon=True).start()
    return RedirectResponse(f"/projects/{project_id}?success=Project+started", status_code=303)


@router.post("/projects/{project_id}/pause")
def pause_project(project_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if project:
        project.status = "paused"
        db.commit()
    return RedirectResponse(f"/projects/{project_id}?success=Project+paused", status_code=303)


@router.post("/projects/{project_id}/resume")
def resume_project(project_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404)
    sub = current_user.subscription
    if not sub or not sub.is_active_plan:
        return RedirectResponse(f"/projects/{project_id}?error=Goi+dich+vu+het+han.+Vui+long+gia+han.", status_code=303)
    project.status = "running"
    db.commit()
    Thread(target=_trigger_pipeline, daemon=True).start()
    return RedirectResponse(f"/projects/{project_id}?success=Project+resumed", status_code=303)


@router.post("/projects/{project_id}/add-keywords")
def add_keywords(
    project_id: int, request: Request,
    keywords_text: str = Form(""),
    db: Session = Depends(get_db),
):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404)

    added = 0
    for kw_text in [k.strip() for k in keywords_text.splitlines() if k.strip()]:
        if not db.query(Keyword).filter(
            Keyword.project_id == project_id, Keyword.keyword == kw_text
        ).first():
            db.add(Keyword(project_id=project_id, keyword=kw_text))
            added += 1
    db.commit()
    return RedirectResponse(f"/projects/{project_id}?success=Added+{added}+keywords", status_code=303)


@router.post("/projects/{project_id}/upload-clusters")
async def upload_clusters(
    project_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Đọc file Excel nhóm từ khóa đã chia sẵn và tạo cluster + article trực tiếp (bỏ qua AI clustering)."""
    import openpyxl

    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404)

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active

        clusters_created = 0
        kw_total = 0
        skipped_rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            cluster_name = str(row[0]).strip()
            if not cluster_name:
                continue

            # Thu thập từ khóa từ cột B trở đi
            keywords = [str(v).strip() for v in row[1:] if v and str(v).strip()]
            if not keywords:
                skipped_rows += 1
                continue

            # Tạo cluster
            cluster = KeywordCluster(
                project_id=project_id,
                cluster_name=cluster_name,
                status="pending",
            )
            db.add(cluster)
            db.flush()

            # Tạo keyword records
            for kw_text in keywords:
                existing = db.query(Keyword).filter(
                    Keyword.project_id == project_id,
                    Keyword.keyword == kw_text,
                ).first()
                if existing:
                    existing.cluster_id = cluster.id
                    existing.status = "clustered"
                else:
                    db.add(Keyword(
                        project_id=project_id,
                        keyword=kw_text,
                        cluster_id=cluster.id,
                        status="clustered",
                    ))
                kw_total += 1

            # Tạo article cho mỗi site
            for ps in project.project_sites:
                db.add(Article(
                    cluster_id=cluster.id,
                    site_id=ps.site_id,
                    project_id=project_id,
                    language=ps.language,
                    status="pending",
                ))

            clusters_created += 1

        db.commit()

        msg = f"Đã tạo {clusters_created} nhóm, {kw_total} từ khóa — chiến dịch bắt đầu viết bài ngay!"
        if skipped_rows:
            msg += f" ({skipped_rows} hàng bỏ qua do thiếu từ khóa)"
        return RedirectResponse(f"/projects/{project_id}?success={msg}", status_code=303)

    except Exception as e:
        logger.error(f"upload_clusters error: {e}")
        return RedirectResponse(f"/projects/{project_id}?error=Lỗi đọc file: {str(e)[:100]}", status_code=303)


@router.get("/projects/{project_id}/edit", response_class=HTMLResponse)
def edit_project_page(project_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404)
    sites = db.query(BlogspotSite).filter(
        BlogspotSite.user_id == current_user.id, BlogspotSite.is_active == True
    ).all()
    selected_site_ids = {ps.site_id for ps in project.project_sites}
    site_languages    = {ps.site_id: ps.language for ps in project.project_sites}
    try:
        current_labels = ", ".join(json.loads(project.custom_labels or "[]"))
    except Exception:
        current_labels = ""
    model_groups = _get_available_models(db, current_user.id)
    return templates.TemplateResponse(request, "project_form.html", {
        "sites": sites, "site_groups": _build_site_groups(sites), "languages": LANGUAGES,
        "models": FREE_MODELS,
        "model_groups": model_groups,
        "default_model": _get_default_model(model_groups, project.ai_model),
        "project": project, "selected_site_ids": selected_site_ids,
        "site_languages": site_languages, "current_labels": current_labels,
        "current_user": current_user, "active_page": "projects",
    })


@router.post("/projects/{project_id}/edit")
def edit_project(
    project_id: int, request: Request,
    name: str = Form(...), description: str = Form(""),
    articles_per_day: int = Form(3), min_interval: int = Form(60), max_interval: int = Form(240),
    ai_model: str = Form("meta-llama/llama-3.1-8b-instruct:free"),
    site_ids: list[int] = Form(default=[]),
    site_languages: list[str] = Form(default=[]),
    backlinks_json: str = Form("[]"), custom_labels_text: str = Form(""),
    db: Session = Depends(get_db),
):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if not project:
        raise HTTPException(status_code=404)

    backlinks, custom_labels = _parse_project_form(backlinks_json, custom_labels_text)
    project.name                = name
    project.description         = description
    project.articles_per_day    = _cap_articles_per_day(articles_per_day, current_user)
    project.min_interval_minutes = min_interval
    project.max_interval_minutes = max_interval
    project.ai_model            = ai_model
    project.backlinks           = json.dumps(backlinks)
    project.custom_labels       = json.dumps(custom_labels, ensure_ascii=False)

    existing     = {ps.site_id: ps for ps in project.project_sites}
    new_site_ids = set(site_ids)
    for site_id, ps in list(existing.items()):
        if site_id not in new_site_ids:
            db.delete(ps)
    for i, site_id in enumerate(site_ids):
        lang = site_languages[i] if i < len(site_languages) else "vi"
        if site_id in existing:
            existing[site_id].language = lang
        else:
            db.add(ProjectSite(project_id=project_id, site_id=site_id, language=lang))

    db.commit()
    return RedirectResponse(f"/projects/{project_id}?success=Da+cap+nhat+du+an", status_code=303)


@router.post("/projects/{project_id}/delete")
def delete_project(project_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = get_current_user(request, db)
    project = db.query(Project).filter(
        Project.id == project_id, Project.user_id == current_user.id
    ).first()
    if project:
        db.delete(project)
        db.commit()
    return RedirectResponse("/projects?success=Project+deleted", status_code=303)


@router.get("/billing", response_class=HTMLResponse)
def billing_page(request: Request, db: Session = Depends(get_db)):
    import json as _json
    from ..models import PLAN_LIMITS
    from ..services.sepay import generate_reference, build_qr_url, expected_amount, MONTH_OPTIONS
    current_user = get_current_user(request, db)

    bank_name   = get_setting(db, "payment_bank_name") or ""
    acct_number = get_setting(db, "payment_account_number") or ""
    acct_holder = get_setting(db, "payment_account_holder") or ""

    payment_config = {
        "bank_name":      bank_name,
        "account_number": acct_number,
        "account_holder": acct_holder,
        "transfer_note":  get_setting(db, "payment_transfer_note") or "",
    }

    sepay_configured = bool(get_setting(db, "sepay_api_key"))
    vn_bank_ready    = bool(bank_name and acct_number)

    uid = current_user.id
    # vn_payment[plan][months] = {reference, amount, qr_url}
    vn_payment = {}
    if vn_bank_ready:
        for plan in ("pro", "business"):
            vn_payment[plan] = {}
            for months in MONTH_OPTIONS:
                ref = generate_reference(uid, plan, months)
                amt = expected_amount(plan, months)
                vn_payment[plan][months] = {
                    "reference": ref,
                    "amount":    amt,
                    "qr_url":    build_qr_url(acct_number, bank_name, amt, ref),
                }

    ls_configured = bool(
        get_setting(db, "ls_api_key") and
        get_setting(db, "ls_store_id") and
        get_setting(db, "ls_pro_variant_id")
    )
    return templates.TemplateResponse(request, "billing.html", {
        "current_user":      current_user,
        "sub":               current_user.subscription,
        "plan_limits":       PLAN_LIMITS,
        "active_page":       "billing",
        "payment_config":    payment_config,
        "vn_bank_ready":     vn_bank_ready,
        "vn_payment":        vn_payment,
        "vn_payment_json":   _json.dumps(vn_payment),
        "month_options":     MONTH_OPTIONS,
        "sepay_configured":  sepay_configured,
        "ls_configured":     ls_configured,
        "ls_pro_price":      get_setting(db, "ls_pro_price") or "$8/month",
        "ls_business_price": get_setting(db, "ls_business_price") or "$20/month",
    })


@router.get("/billing/checkout/{plan}")
def billing_checkout(plan: str, request: Request, db: Session = Depends(get_db)):
    from urllib.parse import quote_plus
    from ..services.lemonsqueezy import create_checkout
    current_user = get_current_user(request, db)

    if plan not in ("pro", "business"):
        return RedirectResponse("/billing?error=Invalid+plan")

    api_key    = get_setting(db, "ls_api_key") or ""
    store_id   = get_setting(db, "ls_store_id") or ""
    variant_id = get_setting(db, f"ls_{plan}_variant_id") or ""

    if not (api_key and store_id and variant_id):
        return RedirectResponse("/billing?error=International+payment+not+configured+yet")

    base_url = str(request.base_url).rstrip("/")
    redirect_url = f"{base_url}/billing?success=payment_completed"
    try:
        checkout_url = create_checkout(
            api_key, store_id, variant_id,
            current_user.email, plan, current_user.id,
            redirect_url,
        )
        return RedirectResponse(checkout_url)
    except ValueError as e:
        return RedirectResponse(f"/billing?error={quote_plus(str(e))}")


@router.post("/billing/webhook/sepay")
async def sepay_webhook(request: Request, db: Session = Depends(get_db)):
    from ..models import User
    from ..services.auth_service import upgrade_plan
    from ..services.sepay import parse_reference, expected_amount

    # Token-based verification: SePay sends ?token=SECRET in webhook URL
    expected_token = get_setting(db, "sepay_webhook_token") or ""
    if expected_token:
        received_token = request.query_params.get("token", "")
        import hmac as _hmac
        if not _hmac.compare_digest(expected_token, received_token):
            logger.warning("SePay webhook: invalid token from %s", request.client.host if request.client else "?")
            return Response(status_code=401)

    try:
        event = await request.json()
    except Exception:
        return Response(status_code=400)

    transfer_type   = event.get("transferType", "")
    transfer_amount = int(event.get("transferAmount", 0))
    content         = event.get("content", "") or ""

    if transfer_type != "in":
        return Response(status_code=200)

    parsed = parse_reference(content)
    if not parsed:
        logger.info("SePay webhook: unrecognized content=%r", content[:100])
        return Response(status_code=200)

    user_id, plan, months = parsed
    required = expected_amount(plan, months)
    if transfer_amount < required:
        logger.warning("SePay webhook: insufficient amount %d < %d for user %d plan=%s months=%d",
                       transfer_amount, required, user_id, plan, months)
        return Response(status_code=200)

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        logger.warning("SePay webhook: user_id=%d not found", user_id)
        return Response(status_code=200)

    upgrade_plan(db, user.id, plan, months=months)
    logger.info("SePay webhook: upgraded user %d (%s) to plan=%s months=%d amount=%d",
                user.id, user.email, plan, months, transfer_amount)
    return Response(content='{"success":true}', media_type="application/json")


@router.post("/billing/webhook/lemonsqueezy")
async def lemonsqueezy_webhook(request: Request, db: Session = Depends(get_db)):
    from ..models import User
    from ..services.auth_service import upgrade_plan
    from ..services.lemonsqueezy import verify_webhook

    payload_bytes = await request.body()
    signature     = request.headers.get("X-Signature", "")
    secret        = get_setting(db, "ls_webhook_secret") or ""

    if not secret:
        logger.error("LemonSqueezy webhook: ls_webhook_secret not configured — rejecting all requests")
        return Response(status_code=503)

    if not verify_webhook(payload_bytes, signature, secret):
        logger.warning("LemonSqueezy webhook: invalid signature from %s",
                       request.client.host if request.client else "?")
        return Response(status_code=401)

    try:
        event = json.loads(payload_bytes)
    except Exception:
        return Response(status_code=400)

    event_name = event.get("meta", {}).get("event_name", "")
    if event_name not in ("order_created", "subscription_created", "subscription_updated"):
        return Response(status_code=200)

    attrs  = event.get("data", {}).get("attributes", {})
    status = attrs.get("status", "")
    if status not in ("paid", "active"):
        return Response(status_code=200)

    custom      = event.get("meta", {}).get("custom_data", {})
    plan        = custom.get("plan", "pro")
    user_id_str = custom.get("user_id", "")
    user_email  = attrs.get("user_email", "")

    user = None
    if user_id_str:
        try:
            user = db.query(User).filter(User.id == int(user_id_str)).first()
        except (ValueError, TypeError):
            pass
    if not user and user_email:
        user = db.query(User).filter(User.email == user_email).first()

    if not user:
        logger.error("LemonSqueezy webhook: could not find user (id=%r, email=%r) for event %s",
                     user_id_str, user_email, event_name)
        return Response(status_code=200)

    upgrade_plan(db, user.id, plan, months=1)
    logger.info("LemonSqueezy webhook: upgraded user %d (%s) to plan=%s event=%s",
                user.id, user.email, plan, event_name)
    return Response(status_code=200)
